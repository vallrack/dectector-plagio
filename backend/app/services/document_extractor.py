import io
import logging

try:
    import fitz # PyMuPDF
except ImportError:
    fitz = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
except ImportError:
    Image = None

import google.generativeai as genai
import os

logger = logging.getLogger(__name__)

class DocumentExtractor:
    """
    Servicio encargado de extraer texto de varios tipos de archivos (documentos y código).
    """

    @staticmethod
    def extract_text(filename: str, content_bytes: bytes) -> str:
        """
        Extrae el texto de un archivo dado su nombre (para inferir extensión) y sus bytes.
        Si es código o texto plano, lo decodifica.
        Si es PDF, Word o Excel, usa librerías específicas.
        Retorna el texto extraído o None si no es soportado/falla.
        """
        ext = filename.split('.')[-1].lower() if '.' in filename else ''
        
        # Archivos que sabemos seguro que son binarios y tienen extracción específica
        if ext == 'pdf':
            return DocumentExtractor._extract_pdf(content_bytes)
        elif ext in ['docx', 'doc']:
            return DocumentExtractor._extract_docx(content_bytes)
        elif ext in ['xlsx', 'xls']:
            return DocumentExtractor._extract_xlsx(content_bytes)
        elif ext in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
            return DocumentExtractor._extract_image(content_bytes)
        
        # Archivos binarios ignorados por defecto (audios, compilados)
        ignored_extensions = {'mp3', 'mp4', 'exe', 'dll', 'zip', 'tar', 'gz', 'rar'}
        if ext in ignored_extensions:
            return None

        # Para cualquier otro archivo (asumimos que es código o texto plano, ej: .py, .java, .txt, .sql, o lenguajes raros)
        try:
            return content_bytes.decode('utf-8')
        except UnicodeDecodeError:
            # Si falla la decodificación, probablemente sea un binario no reconocido, lo ignoramos.
            logger.warning(f"No se pudo decodificar el archivo {filename} como utf-8.")
            return None

    @staticmethod
    def _extract_pdf(content_bytes: bytes) -> str:
        if not fitz:
            logger.error("PyMuPDF (fitz) no está instalado.")
            return None
        
        try:
            text = ""
            # Cargar PDF desde memoria
            with fitz.open(stream=content_bytes, filetype="pdf") as doc:
                for page in doc:
                    text += page.get_text() + "\n"
            return text.strip()
        except Exception as e:
            logger.error(f"Error extrayendo texto de PDF: {e}")
            return None

    @staticmethod
    def _extract_docx(content_bytes: bytes) -> str:
        if not Document:
            logger.error("python-docx no está instalado.")
            return None
        
        try:
            doc = Document(io.BytesIO(content_bytes))
            text = "\n".join([para.text for para in doc.paragraphs])
            return text.strip()
        except Exception as e:
            logger.error(f"Error extrayendo texto de DOCX: {e}")
            return None

    @staticmethod
    def _extract_xlsx(content_bytes: bytes) -> str:
        if not load_workbook:
            logger.error("openpyxl no está instalado.")
            return None
        
        try:
            # data_only=True para leer los valores, no las fórmulas
            wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
            text_lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    # Filtrar celdas vacías y unir con espacio
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        text_lines.append(row_text)
            return "\n".join(text_lines)
        except Exception as e:
            logger.error(f"Error extrayendo texto de XLSX: {e}")
            return None

    @staticmethod
    def _extract_image(content_bytes: bytes) -> str:
        """
        Extrae código o texto de una imagen usando Gemini 1.5 Flash (Vision).
        """
        if not Image:
            logger.error("Pillow (PIL) no está instalado para procesar imágenes.")
            return None

        api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_GENAI_API_KEY")
        if not api_key:
            logger.error("No hay API KEY de Gemini configurada para leer imágenes.")
            return None

        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-1.5-flash')
            image = Image.open(io.BytesIO(content_bytes))
            
            prompt = (
                "Extrae todo el texto y código que veas en esta imagen. "
                "Responde únicamente con el texto extraído, sin formato markdown adicional ni explicaciones."
            )
            response = model.generate_content([prompt, image])
            return response.text
        except Exception as e:
            logger.error(f"Error extrayendo texto de imagen con Gemini: {e}")
            return None
